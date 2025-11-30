from rest_framework import generics, status, views
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.views import APIView
from django.contrib.auth import get_user_model

from .models import CustomUser
from .serializers import (
    UserSerializer, UserCreateSerializer, UserUpdateSerializer,
    LoginSerializer, ChangePasswordSerializer,
    EmployeeListSerializer, SupervisorListSerializer
)
from .permissions import (
    IsAdmin, IsSupervisor, IsAdminOrSupervisor,
    CanManageEmployees, CanManageSupervisors,
    IsOwnerOrSupervisor
)

User = get_user_model()


class LoginView(views.APIView):
    """
    Vista para login de usuarios.
    POST /api/auth/login/
    """
    permission_classes = [AllowAny]
    serializer_class = LoginSerializer
    
    def post(self, request):
        serializer = self.serializer_class(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        
        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)
        
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'user': UserSerializer(user).data
        }, status=status.HTTP_200_OK)


class LogoutView(views.APIView):
    """
    Vista para logout de usuarios.
    POST /api/auth/logout/
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            refresh_token = request.data.get('refresh')
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()
            return Response({'detail': 'Sesión cerrada exitosamente'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'detail': 'Error al cerrar sesión'}, status=status.HTTP_400_BAD_REQUEST)


class ChangePasswordView(views.APIView):
    """
    Vista para cambiar contraseña del usuario actual.
    POST /api/auth/change-password/
    """
    permission_classes = [IsAuthenticated]
    serializer_class = ChangePasswordSerializer
    
    def post(self, request):
        serializer = self.serializer_class(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'detail': 'Contraseña cambiada exitosamente'}, status=status.HTTP_200_OK)


class CurrentUserView(generics.RetrieveUpdateAPIView):
    """
    Vista para obtener y actualizar el perfil del usuario actual.
    GET/PUT /api/auth/me/
    """
    permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer
    
    def get_object(self):
        return self.request.user
    
    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return UserUpdateSerializer
        return UserSerializer


# ==================== VISTAS PARA ADMINISTRADOR ====================
# NOTA: Los supervisores son las cuentas de las empresas.
# Los admins pueden crear y gestionar estas cuentas desde aquí.
# Cada empresa = 1 supervisor activo.

class SupervisorListCreateView(generics.ListCreateAPIView):
    """
    Vista para listar y crear supervisores/cuentas de empresa (solo Admin).
    GET/POST /api/admin/supervisors/
    """
    permission_classes = [CanManageSupervisors]
    
    def get_queryset(self):
        return User.objects.filter(role='supervisor').select_related('company').order_by('-created_at')
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return UserCreateSerializer
        return SupervisorListSerializer
    
    def perform_create(self, serializer):
        # Crear supervisor (cuenta de empresa)
        # Si el admin que crea tiene `company` asignada, usarla como empresa del supervisor.
        # Esto permite que el frontend no tenga que enviar `company` en el formulario.
        company = getattr(self.request.user, 'company', None)
        if company:
            serializer.save(role='supervisor', company=company)
        else:
            # Si no hay company en el usuario, delegar a serializer (y a sus validaciones)
            serializer.save(role='supervisor')


class SupervisorDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Vista para ver, actualizar y eliminar supervisores/cuentas de empresa (solo Admin).
    GET/PUT/DELETE /api/admin/supervisors/{id}/
    """
    permission_classes = [CanManageSupervisors]
    
    def get_queryset(self):
        return User.objects.filter(role='supervisor').select_related('company')
    
    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return UserUpdateSerializer
        return UserSerializer
    
    def perform_destroy(self, instance):
        # Soft delete
        instance.is_active = False
        instance.save()


class AdminStatsView(views.APIView):
    """
    Vista para estadísticas generales del sistema (solo Admin).
    GET /api/admin/stats/
    """
    permission_classes = [IsAdmin]
    
    def get(self, request):
        stats = {
            'total_supervisors': User.objects.filter(role='supervisor', is_active=True).count(),
            'total_employees': User.objects.filter(role='employee', is_active=True).count(),
            'active_users': User.objects.filter(is_active=True).count(),
            'inactive_users': User.objects.filter(is_active=False).count(),
        }
        return Response(stats, status=status.HTTP_200_OK)


# ==================== VISTAS PARA SUPERVISOR ====================

class EmployeeListCreateView(generics.ListCreateAPIView):
    """
    Vista para listar y crear empleados (Supervisor).
    GET/POST /api/supervisor/employees/
    """
    permission_classes = [CanManageEmployees]
    pagination_class = None  # Deshabilitar paginación para esta vista
    
    def get_queryset(self):
        user = self.request.user
        if user.is_admin():
            # Admin ve todos los empleados de todas las empresas
            return User.objects.filter(role='employee').select_related('supervisor', 'company')
        elif user.is_supervisor():
            # Supervisor ve TODOS los empleados de su empresa
            return User.objects.filter(
                role='employee',
                company=user.company
            ).select_related('supervisor', 'company')
        return User.objects.none()
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return UserCreateSerializer
        return EmployeeListSerializer
    
    def perform_create(self, serializer):
        user = self.request.user
        # Obtener el supervisor activo de la empresa
        supervisor = None
        if user.is_supervisor():
            # Si el usuario es supervisor, él mismo es el supervisor
            supervisor = user
        elif user.is_admin() and user.company:
            # Si es admin con empresa, buscar el supervisor de esa empresa
            supervisor = User.objects.filter(
                company=user.company,
                role='supervisor',
                is_active=True
            ).first()
        
        # Asignar el supervisor y la empresa al empleado
        serializer.save(
            supervisor=supervisor,
            company=user.company,
            role='employee'
        )


class EmployeeDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Vista para ver, actualizar y eliminar empleados (Supervisor).
    GET/PUT/DELETE /api/supervisor/employees/{id}/
    """
    permission_classes = [CanManageEmployees]
    pagination_class = None  # Deshabilitar paginación
    
    def get_queryset(self):
        user = self.request.user
        if user.is_admin():
            return User.objects.filter(role='employee').select_related('supervisor', 'company')
        elif user.is_supervisor():
            # Supervisor ve TODOS los empleados de su empresa
            return User.objects.filter(
                role='employee',
                company=user.company
            ).select_related('supervisor', 'company')
        return User.objects.none()
    
    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return UserUpdateSerializer
        # Para GET (detalle), usar EmployeeListSerializer para incluir todos los campos
        from .serializers import EmployeeListSerializer
        return EmployeeListSerializer
    
    def perform_destroy(self, instance):
        # Verificar que el empleado pertenezca a la empresa del supervisor
        if self.request.user.is_supervisor() and instance.company != self.request.user.company:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("No puedes eliminar empleados de otra empresa")
        # Soft delete
        instance.is_active = False
        instance.save()


# ==================== VISTAS PARA EMPLEADO ====================

class EmployeeProfileView(generics.RetrieveAPIView):
    """
    Vista para que el empleado vea su perfil.
    GET /api/employee/me/
    """
    permission_classes = [IsAuthenticated]
    from .serializers import EmployeeListSerializer
    serializer_class = EmployeeListSerializer
    
    def get_object(self):
        return self.request.user


class UploadAvatarView(APIView):
    """
    Vista para subir/actualizar la foto de perfil del usuario autenticado.
    POST /api/auth/upload-avatar/
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        from rest_framework.parsers import MultiPartParser, FormParser
        
        user = request.user
        avatar = request.FILES.get('avatar')
        
        if not avatar:
            return Response(
                {'error': 'No se proporcionó ninguna imagen'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar tipo de archivo
        allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        if avatar.content_type not in allowed_types:
            return Response(
                {'error': 'Tipo de archivo no permitido. Use JPG, PNG, GIF o WEBP.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar tamaño (máx 5MB)
        if avatar.size > 5 * 1024 * 1024:
            return Response(
                {'error': 'La imagen no puede superar los 5MB'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Eliminar avatar anterior si existe
        if user.avatar:
            user.avatar.delete(save=False)
        
        # Guardar nuevo avatar
        user.avatar = avatar
        user.save()
        
        # Construir URL del avatar
        avatar_url = request.build_absolute_uri(user.avatar.url) if user.avatar else None
        
        return Response({
            'message': 'Foto de perfil actualizada exitosamente',
            'avatar_url': avatar_url
        }, status=status.HTTP_200_OK)
    
    def delete(self, request):
        """Eliminar foto de perfil"""
        user = request.user
        
        if user.avatar:
            user.avatar.delete(save=False)
            user.avatar = None
            user.save()
            return Response({'message': 'Foto de perfil eliminada'}, status=status.HTTP_200_OK)
        
        return Response({'error': 'No hay foto de perfil para eliminar'}, status=status.HTTP_400_BAD_REQUEST)


class EmployeeExportDataView(APIView):
    """
    Vista para que el empleado descargue todos sus datos en Excel.
    GET /api/auth/employee/export-my-data/
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from apps.sensors.models import ProcessedMetrics
        from apps.analytics.models import FatigueAlert, RoutineRecommendation
        
        user = request.user
        
        # Crear workbook
        wb = Workbook()
        
        # === HOJA 1: INFORMACIÓN PERSONAL ===
        ws_personal = wb.active
        ws_personal.title = "Información Personal"
        
        # Encabezado
        ws_personal['A1'] = 'INFORMACIÓN PERSONAL Y LABORAL'
        ws_personal['A1'].font = Font(bold=True, size=14)
        ws_personal['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_personal['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_personal.merge_cells('A1:B1')
        
        # Datos personales
        personal_data = [
            ['Campo', 'Valor'],
            ['Nombre Completo', user.get_full_name()],
            ['Email', user.email],
            ['Teléfono', user.phone or 'No registrado'],
            ['Departamento', user.department or 'No asignado'],
            ['Puesto', user.position or 'No asignado'],
            ['Rol', user.get_role_display()],
            ['Empresa', user.company.name if user.company else 'No asignada'],
            ['Supervisor', user.supervisor.get_full_name() if user.supervisor else 'Sin supervisor'],
            ['Estado', 'Activo' if user.is_active else 'Inactivo'],
            ['Fecha de Registro', user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else 'N/A'],
        ]
        
        for row_idx, row_data in enumerate(personal_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_personal.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Encabezados de columna
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Ajustar ancho de columnas
        ws_personal.column_dimensions['A'].width = 25
        ws_personal.column_dimensions['B'].width = 40
        
        # === HOJA 2: HISTORIAL DE MÉTRICAS ===
        ws_metrics = wb.create_sheet("Historial de Métricas")
        ws_metrics['A1'] = 'HISTORIAL DE MÉTRICAS'
        ws_metrics['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_metrics['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws_metrics.merge_cells('A1:H1')
        
        # Obtener métricas del empleado
        metrics = ProcessedMetrics.objects.filter(employee=user).order_by('-window_start')[:100]
        
        if metrics.exists():
            headers = ['Fecha', 'HR Promedio', 'HR Máximo', 'SpO2 Promedio', 'HRV RMSSD', 'Índice Fatiga', 'Nivel Actividad', 'Estado']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_metrics.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for row_idx, metric in enumerate(metrics, start=4):
                ws_metrics.cell(row=row_idx, column=1, value=metric.window_start.strftime('%Y-%m-%d %H:%M:%S'))
                ws_metrics.cell(row=row_idx, column=2, value=f"{metric.hr_avg:.1f}")
                ws_metrics.cell(row=row_idx, column=3, value=f"{metric.hr_max:.1f}")
                ws_metrics.cell(row=row_idx, column=4, value=f"{metric.spo2_avg:.1f}")
                ws_metrics.cell(row=row_idx, column=5, value=f"{metric.hrv_rmssd:.1f}" if metric.hrv_rmssd else "N/A")
                ws_metrics.cell(row=row_idx, column=6, value=f"{metric.fatigue_index:.2f}" if metric.fatigue_index else "N/A")
                ws_metrics.cell(row=row_idx, column=7, value=metric.get_activity_level_display() if hasattr(metric, 'activity_level') else "N/A")
                ws_metrics.cell(row=row_idx, column=8, value="Normal")
        else:
            ws_metrics.cell(row=3, column=1, value="Sin datos registrados aún")
            ws_metrics.cell(row=3, column=1).font = Font(italic=True, color="999999")
        
        # Ajustar anchos
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws_metrics.column_dimensions[col].width = 18
        
        # === HOJA 3: ALERTAS RECIBIDAS ===
        ws_alerts = wb.create_sheet("Alertas Recibidas")
        ws_alerts['A1'] = 'ALERTAS RECIBIDAS'
        ws_alerts['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_alerts['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        ws_alerts.merge_cells('A1:F1')
        
        # Obtener alertas del empleado
        alerts = FatigueAlert.objects.filter(employee=user).order_by('-timestamp')
        
        if alerts.exists():
            headers = ['Fecha', 'Tipo', 'Severidad', 'Descripción', 'Estado', 'Resuelta']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_alerts.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for row_idx, alert in enumerate(alerts, start=4):
                ws_alerts.cell(row=row_idx, column=1, value=alert.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                ws_alerts.cell(row=row_idx, column=2, value=alert.alert_type)
                ws_alerts.cell(row=row_idx, column=3, value=alert.get_severity_display())
                ws_alerts.cell(row=row_idx, column=4, value=alert.message)
                ws_alerts.cell(row=row_idx, column=5, value="Resuelta" if alert.resolved else "Pendiente")
                ws_alerts.cell(row=row_idx, column=6, value=alert.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if alert.resolved_at else "N/A")
        else:
            ws_alerts.cell(row=3, column=1, value="Sin datos registrados aún")
            ws_alerts.cell(row=3, column=1).font = Font(italic=True, color="999999")
        
        # Ajustar anchos
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_alerts.column_dimensions[col].width = 20
        ws_alerts.column_dimensions['D'].width = 50
        
        # === HOJA 4: RECOMENDACIONES APLICADAS ===
        ws_recs = wb.create_sheet("Recomendaciones")
        ws_recs['A1'] = 'RECOMENDACIONES APLICADAS'
        ws_recs['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_recs['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws_recs.merge_cells('A1:E1')
        
        # Obtener recomendaciones del empleado
        recommendations = RoutineRecommendation.objects.filter(employee=user).order_by('-created_at')
        
        if recommendations.exists():
            headers = ['Fecha', 'Tipo', 'Descripción', 'Estado', 'Aplicada']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_recs.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for row_idx, rec in enumerate(recommendations, start=4):
                ws_recs.cell(row=row_idx, column=1, value=rec.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                ws_recs.cell(row=row_idx, column=2, value=rec.get_recommendation_type_display())
                ws_recs.cell(row=row_idx, column=3, value=rec.description)
                ws_recs.cell(row=row_idx, column=4, value=rec.get_status_display())
                ws_recs.cell(row=row_idx, column=5, value="Sí" if rec.applied else "No")
        else:
            ws_recs.cell(row=3, column=1, value="Sin datos registrados aún")
            ws_recs.cell(row=3, column=1).font = Font(italic=True, color="999999")
        
        # Ajustar anchos
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_recs.column_dimensions[col].width = 20
        ws_recs.column_dimensions['C'].width = 60
        
        # Guardar en respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="mis_datos_{user.id}.xlsx"'
        
        wb.save(response)
        return response

